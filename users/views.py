from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.exceptions import AuthenticationFailed
from rest_framework import status
from rest_framework.permissions import BasePermission
from .models import CustomUser
from .serializers import UserSerializer
import jwt
from rest_framework.generics import RetrieveAPIView
import datetime
import random
from django.core.mail import send_mail
from rest_framework import viewsets
from .models import PlageHoraire, Semaine, JourSemaine, Enseignant
from .serializers import PlageHoraireSerializer, SemaineSerializer, JourSemaineSerializer, EnseignantSerializer
from .models import Matiere, Groupe,  ChargeHebdomadaire, AffectationEnseignant
from .serializers import MatiereSerializer, GroupeSerializer, ChargeHebdomadaireSerializer, AffectationEnseignantSerializer
from django.http import JsonResponse, HttpResponse
from ortools.sat.python import cp_model
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from .models import EmploiDuTemps, Groupe, Enseignant, Matiere, PlageHoraire, Salle, AffectationEnseignant, ChargeHebdomadaire, Semaine, JourSemaine
class IsAuthenticated(BasePermission):
    def has_permission(self, request, view):
        token = request.headers.get('Authorization')

        if not token:
            raise AuthenticationFailed('Token not provided', code='token_not_provided')

        try:
            payload = jwt.decode(token, 'secret', algorithms=['HS256'])
        except jwt.ExpiredSignatureError:
            raise AuthenticationFailed('Token has expired', code='token_expired')

        user = CustomUser.objects.filter(id_u=payload['id']).first()
        if user is None:
            raise AuthenticationFailed('User not found!', code='user_not_found')

        return True


class VerifierEmailView(APIView):
    def post(self, request):
        email = request.data.get('email')
        verification_code = ''.join(random.choices('0123456789', k=6))
        send_mail(
            'Email Verification Code',
            f'Your email verification code is: {verification_code}',
            '22034@supnum.mr',
            [email],
            fail_silently=False,
        )
        return Response({"verification_code": verification_code}, status=status.HTTP_200_OK)


class RegisterView(APIView):
    def post(self, request):
        serializer = UserSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
class LoginView(APIView):
    def post(self, request):
        login_or_email = request.data.get('login_or_email')
        password = request.data.get('password')

        if '@' in login_or_email:
            user = CustomUser.objects.filter(email=login_or_email).first()
        else:
            user = CustomUser.objects.filter(login=login_or_email).first()

        if user is None:
            raise AuthenticationFailed('User not found!', code='user_not_found')

        if not user.check_password(password):
            raise AuthenticationFailed('Incorrect password!', code='incorrect_password')

        payload = {
            'id': user.id_u,
            'exp': datetime.datetime.utcnow() + datetime.timedelta(hours=6),
            'iat': datetime.datetime.utcnow()
        }
        token = jwt.encode(payload, 'secret', algorithm='HS256')

        response = Response({'jwt': token,'role':user.role,'id_u':user.id_u})
        response.set_cookie(key='jwt', value=token, httponly=True, secure=False, samesite='Lax')
        return response


class UserView(APIView):
    def get(self, request, id_u):
        user = CustomUser.objects.filter(id_u=id_u).first()

        if not user:
            return Response({'error': 'User not found!'}, status=404)

        serializer = UserSerializer(user)
        return Response(serializer.data)

class LogoutView(APIView):
    def post(self, request):
        response = Response({'message': 'success'})
        response.delete_cookie('jwt')
        return response
    
    


class PlageHoraireViewSet(viewsets.ModelViewSet):
    queryset = PlageHoraire.objects.all()
    serializer_class = PlageHoraireSerializer

class SemaineViewSet(viewsets.ModelViewSet):
    queryset = Semaine.objects.all()
    serializer_class = SemaineSerializer

class JourSemaineViewSet(viewsets.ModelViewSet):
    queryset = JourSemaine.objects.all()
    serializer_class = JourSemaineSerializer

class EnseignantViewSet(viewsets.ModelViewSet):
    queryset = Enseignant.objects.all()
    # lookup_field = 'identifiant'
    serializer_class = EnseignantSerializer
    # def update(self, request, *args, **kwargs):
    #     """Gestion des mises à jour avec PUT."""
    #     return super().update(request, *args, **kwargs)

class EnseignantByIdentifiantView(RetrieveAPIView):
    serializer_class = EnseignantSerializer
    lookup_field = 'identifiant'

    def get_queryset(self):
        return Enseignant.objects.all()
class MatiereViewSet(viewsets.ModelViewSet):
    queryset = Matiere.objects.all()
    serializer_class = MatiereSerializer


class GroupeViewSet(viewsets.ModelViewSet):
    queryset = Groupe.objects.all()
    serializer_class = GroupeSerializer



class ChargeHebdomadaireViewSet(viewsets.ModelViewSet):
    queryset = ChargeHebdomadaire.objects.all()
    serializer_class = ChargeHebdomadaireSerializer


class AffectationEnseignantViewSet(viewsets.ModelViewSet):
    queryset = AffectationEnseignant.objects.all()
    serializer_class = AffectationEnseignantSerializer



from ortools.sat.python import cp_model
from django.http import JsonResponse
from django.utils.timezone import now
from .models import (
    Semaine, JourSemaine, Groupe, Enseignant, Matiere, PlageHoraire, 
    Salle, AffectationEnseignant, ChargeHebdomadaire, EmploiDuTemps, Filiere
)

def generer_emploi_du_temps_filiere(request, filiere_id):
    model = cp_model.CpModel()

    semaine_actuelle = Semaine.objects.latest('date_debut')
    jours_semaine = list(JourSemaine.objects.filter(semaine=semaine_actuelle))
    
    filiere = Filiere.objects.get(id=filiere_id)
    groupes = list(Groupe.objects.filter(filiere=filiere))
    matieres = list(Matiere.objects.filter(filiere=filiere))
    enseignants = list(Enseignant.objects.filter(affectationenseignant__matiere__in=matieres).distinct())
    plages = list(PlageHoraire.objects.all())
    salles = list(Salle.objects.all())
    affectations = list(AffectationEnseignant.objects.filter(matiere__in=matieres))

    x = {}
    for g in groupes:
        for e in enseignants:
            for m in matieres:
                for p in plages:
                    for j in jours_semaine:
                        x[g.id, e.id_Es, m.id, p.id_plh, j.id_jrs] = model.NewBoolVar(f"x_{g.id}_{e.id_Es}_{m.id}_{p.id_plh}_{j.id_jrs}")


    for charge in ChargeHebdomadaire.objects.filter(matiere__in=matieres):
        total_heures = charge.cm_heures + charge.td_heures + charge.tp_heures
        model.Add(
            sum(x[charge.groupe.id, e.id_Es, charge.matiere.id, p.id_plh, j.id_jrs] 
                for e in enseignants for p in plages for j in jours_semaine
                if AffectationEnseignant.objects.filter(enseignant=e, matiere=charge.matiere).exists()
            ) == total_heures
        )


    for g in groupes:
        for p in plages:
            for j in jours_semaine:
                model.Add(
                    sum(x[g.id, e.id_Es, m.id, p.id_plh, j.id_jrs] 
                        for e in enseignants for m in matieres
                    ) <= 1
                )


    for e in enseignants:
        for p in plages:
            for j in jours_semaine:
                model.Add(
                    sum(x[g.id, e.id_Es, m.id, p.id_plh, j.id_jrs] 
                        for g in groupes for m in matieres
                    ) <= 1
                )

    for s in salles:
        for p in plages:
            for j in jours_semaine:
                model.Add(
                    sum(x[g.id, e.id_Es, m.id, p.id_plh, j.id_jrs] 
                        for g in groupes for e in enseignants for m in matieres
                        if AffectationEnseignant.objects.filter(enseignant=e, matiere=m).exists()
                    ) <= 1 
                )

    model.Maximize(
        sum(x[g.id, e.id_Es, m.id, p.id_plh, j.id_jrs] 
            for g in groupes for e in enseignants for m in matieres for p in plages for j in jours_semaine
        )
    )


    solver = cp_model.CpSolver()
    status = solver.Solve(model)

    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        emplois = []
        for g in groupes:
            for e in enseignants:
                for m in matieres:
                    for p in plages:
                        for j in jours_semaine:
                            if solver.Value(x[g.id, e.id_Es, m.id, p.id_plh, j.id_jrs]) == 1:
                                salle_disponible = next(
                                    (s for s in salles if not EmploiDuTemps.objects.filter(plage_horaire=p, jour=j, salle=s).exists()), 
                                    None
                                )

                                emploi = EmploiDuTemps(
                                    groupe=g,
                                    enseignant=e,
                                    matiere=m,
                                    plage_horaire=p,
                                    salle=salle_disponible,
                                    jour=j,
                                    heure_debut_cours=p.heure_debut,
                                    heure_fin_cours=p.heure_fin
                                )
                                emploi.save()
                                emplois.append({
                                    "groupe": g.nom,
                                    "enseignant": e.nom,
                                    "matiere": m.nom,
                                    "plage_horaire": f"{p.heure_debut} - {p.heure_fin}",
                                    "date_jour": j.date_jour.strftime('%Y-%m-%d'),
                                    "salle": salle_disponible.nom if salle_disponible else "Aucune salle disponible"
                                })
        return JsonResponse({"status": "success", "message": "Emploi du temps généré", "data": emplois})
    else:
        return JsonResponse({"status": "error", "message": "Aucune solution trouvée"})


import pandas as pd
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.pdfgen import canvas
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict

def telecharger_emploi_excel(request, groupe_id):
    emplois = EmploiDuTemps.objects.filter(groupe__id=groupe_id).order_by('jour__date_jour', 'heure_debut_cours')

    data = defaultdict(list)
    for e in emplois:
        data[e.jour.date_jour.strftime('%A %d/%m/%Y')].append([
            f"{e.heure_debut_cours} - {e.heure_fin_cours}",
            e.matiere.nom,
            e.enseignant.nom,
            e.salle.nom if e.salle else ""
        ])

    wb = Workbook()
    ws = wb.active
    ws.title = "Emploi du Temps"

 
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    ws.append(["Jour", "Horaire", "Matière", "Enseignant", "Salle"])
    for col in ws[1]:
        col.fill = header_fill
        col.font = header_font
        col.border = border
        col.alignment = Alignment(horizontal="center")

 
    for jour, cours in data.items():
        for i, row in enumerate(cours):
            ws.append([jour if i == 0 else "", *row])  

  
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="emploi_du_temps.xlsx"'
    wb.save(response)
    return response

def telecharger_emploi_pdf(request, groupe_id):
    emplois = EmploiDuTemps.objects.filter(groupe__id=groupe_id).order_by('jour__date_jour', 'heure_debut_cours')

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []

    data = [["Jour", "Horaire", "Matière", "Enseignant", "Salle"]]
    for e in emplois:
        data.append([
            e.jour.date_jour.strftime('%A %d/%m/%Y'),
            f"{e.heure_debut_cours} - {e.heure_fin_cours}",
            e.matiere.nom,
            e.enseignant.nom,
            e.salle.nom if e.salle else ""
        ])


    table = Table(data, colWidths=[120, 120, 180, 180, 100])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(table)
    doc.build(elements)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="emploi_du_temps.pdf"'
    response.write(buffer.getvalue())
    buffer.close()
    return response

